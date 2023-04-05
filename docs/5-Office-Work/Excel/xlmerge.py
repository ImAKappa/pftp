"""xlmerge.py

Merge separate Excel file into one
"""

from openpyxl import load_workbook
from openpyxl import Workbook
from pathlib import Path

import openpyxl
from copy import copy


def copy_sheet(source_sheet, target_sheet):
    copy_cells(source_sheet, target_sheet)  # copy all the cel values and styles
    copy_sheet_attributes(source_sheet, target_sheet)


def copy_sheet_attributes(source_sheet, target_sheet):
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.merged_cells = copy(source_sheet.merged_cells)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.freeze_panes = copy(source_sheet.freeze_panes)

    # set row dimensions
    # So you cannot copy the row_dimensions attribute. Does not work (because of meta data in the attribute I think). So we copy every row's row_dimensions. That seems to work.
    for rn in range(len(source_sheet.row_dimensions)):
        target_sheet.row_dimensions[rn] = copy(source_sheet.row_dimensions[rn])

    if source_sheet.sheet_format.defaultColWidth is None:
        print("Unable to copy default column wide")
    else:
        target_sheet.sheet_format.defaultColWidth = copy(
            source_sheet.sheet_format.defaultColWidth
        )

    # set specific column width and hidden property
    # we cannot copy the entire column_dimensions attribute so we copy selected attributes
    for key, value in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[key].min = copy(
            source_sheet.column_dimensions[key].min
        )  # Excel actually groups multiple columns under 1 key. Use the min max attribute to also group the columns in the targetSheet
        target_sheet.column_dimensions[key].max = copy(
            source_sheet.column_dimensions[key].max
        )  # https://stackoverflow.com/questions/36417278/openpyxl-can-not-read-consecutive-hidden-columns discussed the issue. Note that this is also the case for the width, not onl;y the hidden property
        target_sheet.column_dimensions[key].width = copy(
            source_sheet.column_dimensions[key].width
        )  # set width for every column
        target_sheet.column_dimensions[key].hidden = copy(
            source_sheet.column_dimensions[key].hidden
        )


def copy_cells(source_sheet, target_sheet):
    for (row, col), source_cell in source_sheet._cells.items():
        target_cell = target_sheet.cell(column=col, row=row)

        target_cell._value = source_cell._value
        target_cell.data_type = source_cell.data_type

        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.border = copy(source_cell.border)
            target_cell.fill = copy(source_cell.fill)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.hyperlink:
            target_cell._hyperlink = copy(source_cell.hyperlink)

        if source_cell.comment:
            target_cell.comment = copy(source_cell.comment)


EXCEL_SHEET_NAME_CHAR_LIMIT = 31

if __name__ == "__main__":
    input_dir = Path("docs/5-Office-Work/Excel/data")
    output_file = Path("combined.xlsx")
    wb_target = openpyxl.Workbook()

    for xl_path in input_dir.iterdir():
        sheet_name = xl_path.stem[: min(EXCEL_SHEET_NAME_CHAR_LIMIT, len(xl_path.stem))]
        target_sheet = wb_target.create_sheet(sheet_name)
        wb_source = openpyxl.load_workbook(xl_path, data_only=True)
        source_sheet = wb_source.active

        copy_sheet(source_sheet, target_sheet)

        if "Sheet" in wb_target.sheetnames:  # remove default sheet
            wb_target.remove(wb_target["Sheet"])

    wb_target.save(output_file)